import json
import math
import os
import sys
import pandas as pd
import warnings

from openpyxl import load_workbook
from openpyxl.styles import numbers
from pathlib import Path
from datetime import datetime, date

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

home = os.path.expanduser('~')

def main(prod):

    datum = datetime.today().strftime("%Y%m%d %H%M%S")

    print(datum + ": Start der Verarbeitung")

    my_dir = Path(__file__).parent

    # Konfiguration einlesen
    f = open(my_dir / 'config_v2.json')
    data = json.load(f)
    data['Prod']=prod
    f.close()

    # Aufruf der Adressmaster-Datei
    alle_projekte = adressmaster_einlesen(data)

    df_pmd = pmd_einlesen(data)

    basis_excel_erstellen(data, alle_projekte, df_pmd)

    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Ende der Verarbeitung")

def adressmaster_einlesen(data):

    def adressmaster_einlesen(data, lk, alle_projekte):
        datum = datetime.today().strftime("%Y%m%d %H%M%S")
        print(datum + ": Adressmaster " + lk + " einlesen")

        if data['Prod'] == "j":
            quell_datei = home + "/" + data["oew_ablage_adressen"] + data[lk + "_ADRESS_MASTER"]
        else:
            quell_datei = home + "/" + data["input_dir"] + data["ADRESS_MASTER_TEST"]

        df = pd.read_excel(quell_datei, sheet_name=QUELL_SHEET, header=14)

        alle_projekte = adressmaster_auswerten(alle_projekte, df, lk)

        return alle_projekte

    def adressmaster_auswerten(alle_projekte, df, lk):

        # hgf eAktz ermitteln
        df_gefilter_hgf_untervers = df [df["hgf_untervers"] == 1]
        df_gefiltert = df_gefilter_hgf_untervers.drop_duplicates(subset=['ortsname',
                                                  'gemark_nam',
                                                  'bauabschnitt_code',
                                                  'hgf_aktenzeichnen_quelle'
                                                  ])

        for _, row in df_gefiltert.iterrows():
            alle_projekte.append({
                "landkreis": lk,
                "gemeinde": row["ortsname"],
                "gemarkung": row["gemark_nam"],
                "eaktz": row["hgf_aktenzeichnen_quelle"],
                "programm": "hgf",
                "cluster": row["bau_cluster"],
                "los": row["bau_los"],
                "bc_code": row["bauabschnitt_code"]
            })

        # dgf eAktz ermitteln
        df_gefilter_dgf_untervers = df [df["dgf_untervers"] == 1]
        df_gefiltert = df_gefilter_dgf_untervers.drop_duplicates(subset=['ortsname',
                                                  'gemark_nam',
                                                  'bauabschnitt_code',
                                                  'dgf_aktenzeichnen_quelle'
                                                  ])

        for _, row in df_gefiltert.iterrows():
            alle_projekte.append({
                "landkreis": lk,
                "gemeinde": row["ortsname"],
                "gemarkung": row["gemark_nam"],
                "eaktz": row["dgf_aktenzeichnen_quelle"],
                "programm": "dgf",
                "cluster": row["bau_cluster"],
                "los": row["bau_los"],
                "bc_code": row["bauabschnitt_code"]
            })

        return alle_projekte

    QUELL_SHEET = data["Quellsheet_Adressmaster"]

    alle_projekte = []
    if data['Prod'] == "j":
        #Landkreis ADK einlesen
        lk = "ADK"
        alle_projekte = adressmaster_einlesen(data, lk, alle_projekte)

        #Landkreis FDS einlesen
        lk = "FDS"
        alle_projekte = adressmaster_einlesen(data, lk, alle_projekte)

        # Landkreis LBC
        lk = "BC"
        alle_projekte = adressmaster_einlesen(data, lk, alle_projekte)

        # Landkreis REU
        lk = "REU"
        alle_projekte = adressmaster_einlesen(data, lk, alle_projekte)

        # Landkreis SIG
        lk = "SIG"
        alle_projekte = adressmaster_einlesen(data, lk, alle_projekte)

        # Landkreis ZAK
        lk = "ZAK"
        alle_projekte = adressmaster_einlesen(data, lk, alle_projekte)
    else:
        # Landkreis FDS einlesen
        lk = "FDS"
        alle_projekte = adressmaster_einlesen(data, lk, alle_projekte)

    return alle_projekte

def basis_excel_erstellen(data, alle_projekte, df_pmd):

    def baudaten_ermitteln(df_pmd, bc_code):

            pl_beginn = ""
            pl_ende = ""
            ba_beginn = ""
            ba_ende = ""
            #if bc_code == "FDS_C1_01-01":
            #    print("hallo")
            # Abfrage der Zeilen, die in der Planung sind (GU-Model)
            fixe_kriterien = {
                "PL/BA/PB": "PL"
            }

            df_pmd_gefiltert = df_pmd[df_pmd['BC-Code'] == bc_code]

            for spalte, wert in fixe_kriterien.items():
                df_pmd_gefiltert = df_pmd_gefiltert[df_pmd_gefiltert[spalte].str.strip() == wert]
            if len(df_pmd_gefiltert) == 1:
                pl_beginn = df_pmd_gefiltert["Beginn nach EZA/BV"].iloc[0]
                pl_ende = df_pmd_gefiltert["Beginn nach EZA/BV"].iloc[0]

            # Abfrage der Zeilen, die im Bau sind (GU-Model)
            fixe_kriterien = {
                "PL/BA/PB": "BA"
            }

            df_pmd_gefiltert = df_pmd[df_pmd['BC-Code'] == bc_code]

            for spalte, wert in fixe_kriterien.items():
                df_pmd_gefiltert = df_pmd_gefiltert[df_pmd_gefiltert[spalte].str.strip() == wert]

            if len(df_pmd_gefiltert) == 1:
                ba_beginn = df_pmd_gefiltert["Beginn nach EZA/BV"].iloc[0]
                ba_ende = df_pmd_gefiltert["Beginn nach EZA/BV"].iloc[0]


            # Abfrage der Zeilen, die in Planung und Bau sind (GÜ-Model)
            fixe_kriterien = {
                "PL/BA/PB": "PB"
            }

            df_pmd_gefiltert = df_pmd[df_pmd['BC-Code'] == bc_code]

            for spalte, wert in fixe_kriterien.items():
                df_pmd_gefiltert = df_pmd_gefiltert[df_pmd_gefiltert[spalte].str.strip() == wert]

            if len(df_pmd_gefiltert) == 1:
                pl_beginn = df_pmd_gefiltert["Beginn nach EZA/BV"].iloc[0]
                pl_ende = df_pmd_gefiltert["Beginn nach EZA/BV"].iloc[0]

            return pl_beginn, pl_ende, ba_beginn, ba_ende

    # Umwandlung der Daten in einen Dataframe
    df = pd.DataFrame(alle_projekte)

    # Sortieren der Daten nach Landkreis, Ausschreibung, Gemeinde und Gemarkung
    df_gesamt = df.sort_values(by=['landkreis', 'gemeinde', 'gemarkung', 'programm', 'bc_code'])

    #schluessel = ["landkreis", "gemeinde", "gemarkung", "programm"]

    #hat_code = df_gesamt["bc_code"].notna() & (df_gesamt["bc_code"].astype(str).str.strip() != "")

    #gruppe_hat_code = hat_code.groupby(
    #    [df_gesamt[c] for c in schluessel]
    #).transform("any")

    #df_result = df_gesamt[hat_code | ~gruppe_hat_code]

    # Pfad und Dateiname definieren für die Ausgabe
    template = home + "/" + data['output_dir'] + data['status_template']  #Status_Template
    sheet = "Master"
    datum = datetime.today().strftime("%Y%m%d")
    ausgabe = home + "/" + data['output_dir'] + f"AP24_Master_Status_V2.xlsx"

    column_mapping = {
        "eaktz": 1,
        "landkreis": 2,
        "gemeinde": 3,
        "gemarkung": 4,
        "programm": 5,
        "cluster": 6,
        "los": 7,
        "bc_code": 8
    }

    # Existierende Datei laden (Formatierung & Formeln bleiben erhalten)
    wb = load_workbook(template)
    ws = wb[sheet]  # ggf. Sheet-Name anpassen

    START_ROW = 5  # Zeile, ab der Daten geschrieben werden (1 = erste Zeile)

    #
    for df_col, excel_col in column_mapping.items():
        for row_idx, value in enumerate(df_gesamt[df_col], start=START_ROW):
            ws.cell(row=row_idx, column=excel_col, value=value)
            bc_code = ws.cell(row=row_idx, column=8).value
            #if bc_code != "" and not (isinstance(bc_code, float) and math.isnan(bc_code)):
            if pd.notna(bc_code) and bc_code != "":
                pl_beginn, pl_ende, ba_beginn, ba_ende = baudaten_ermitteln(df_pmd, bc_code)
                ws.cell(row=row_idx, column=9, value=pl_beginn)
                ws.cell(row=row_idx, column=10, value=pl_ende)
                ws.cell(row=row_idx, column=11, value=ba_beginn)
                ws.cell(row=row_idx, column=12, value=ba_ende)


    #        # Datumspalten formatieren
    #        if df_col in ["Planung Beginn", "Planung Ende", "Bauphase Beginn", "Bauphase Ende"]:
    #            cell.number_format = "DD.MM.YYYY"


    wb.save(ausgabe)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Basisdaten in Statusdatei befüllt")

    return ausgabe

def pmd_einlesen(data):

    if data['Prod'] == "j":
        pm_datei = Path(home + "/" + data["input_dir"] + data['pmd_datei_test'])
        #pm_datei = Path(home + "/" + data["oew_ablage_pmd"] + data['pmd_datei'])
    else:
        pm_datei = Path(home + "/" + data["input_dir"] + data['pmd_datei_test'])


    sheet_namen = [
        "ADK",
        "BC",
        "RT",
        "FDS",
        "SIG",
        "ZAK"
    ]

    sheets = pd.read_excel(
        pm_datei,
        sheet_name=sheet_namen
    )

    df = pd.concat(
        [
            df.assign(sheet_name=name)
            for name, df in sheets.items()
        ],
        ignore_index=True
    )

    return df

if __name__ == '__main__':
    prod = sys.argv[1]
    main(prod)