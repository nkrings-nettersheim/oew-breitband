import os
import json
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

home = os.path.expanduser('~')

my_dir = Path(__file__).parent
f = open(my_dir / 'config.json')
data = json.load(f)
input_dir = home + "/" + data['input_dir']
output_dir = home + "/" + data['output_dir']
f.close()

BASIS_DATEI    = input_dir + "AP24_Master_Status.xlsx"
QUELL_DATEI    = input_dir + "2026-03-06_ProjektcontrollingTB.xlsx"
BASIS_SHEET    = "Master"   # Sheet-Name anpassen
QUELL_SHEET    = "Übersicht"   # Sheet-Name anpassen
START_ROW      = 5

# Quelldatei als DataFrame laden (Spalte B = Suchschlüssel)
df_quelle = pd.read_excel(QUELL_DATEI, sheet_name=QUELL_SHEET, header=0)

# Lookup-Dict aufbauen: Wert aus Spalte B → relevante Spalten
# Spalten: C=2, D=3, E=4, I=8, J=9, K=10, L=11 (0-basiert)
lookup = {}
for _, row in df_quelle.iterrows():
    key = row.iloc[1]  # Spalte B
    if pd.notna(key):
        lookup[key] = {
            "C": row.iloc[2],   # Spalte C GÜ
            "D": row.iloc[3],   # Spalte D GU
            "E": row.iloc[4],   # Spalte E Adressen
            #"I": row.iloc[8],   # Spalte I Start Planung
            #"J": row.iloc[9],   # Spalte J Ende Planung
            "K": row.iloc[10],  # Spalte K Start Bau
            "L": row.iloc[11],  # Spalte L Ende Bau
        }

# Zielspalten in der Basis-Excel (1-basiert für openpyxl)
# Passe die Spaltennummern an, wohin die Werte sollen
ziel_spalten = {
    "C": 11,    # Quell-C → Basis-Spalte K
    "D": 12,    # Quell-D → Basis-Spalte L
    "E": 13,    # Quell-E → Basis-Spalte M
    #"I": 9,    # Quell-I → Basis-Spalte I
    #"J": 10,   # Quell-J → Basis-Spalte J
    "K": 28,   # Quell-K → Basis-Spalte K
    "L": 29,   # Quell-L → Basis-Spalte L
}

# Basis-Excel laden und befüllen
wb = load_workbook(BASIS_DATEI)
ws = wb[BASIS_SHEET]

treffer = 0
for row_idx in range(START_ROW, ws.max_row + 1):
    # Suchwert aus Spalte J der Basis-Excel
    suchwert = ws.cell(row=row_idx, column=10).value  # Spalte J = 10

    if suchwert is None:
        continue

    if suchwert in lookup:
        treffer += 1
        for quell_col, ziel_col in ziel_spalten.items():
            ws.cell(row=row_idx, column=ziel_col, value=lookup[suchwert][quell_col])

wb.save(BASIS_DATEI)
print(f"Fertig – {treffer} Treffer gefunden und übertragen.")