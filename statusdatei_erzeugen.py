import json
import os
import pandas as pd
import warnings

from openpyxl import load_workbook
from openpyxl.styles import numbers
from pathlib import Path
from datetime import datetime


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

home = os.path.expanduser('~')

def main():

    #home = os.path.expanduser('~')

    datum = datetime.today().strftime("%Y%m%d %H%M%S")

    print(datum + ": Start der Verarbeitung")

    my_dir = Path(__file__).parent

    # Konfiguration einlesen
    f = open(my_dir / 'config.json')
    data = json.load(f)
    input_dir = home + "/" + data['input_dir']
    output_dir = home + "/" + data['output_dir']
    f.close()

    # Liste der eAkz einlesen
    f = open(my_dir / 'eAktenzeichen.json')
    eAkz = json.load(f)
    f.close()
    eAkz_dictionary = {eintrag["aktenzeichen"]: eintrag for eintrag in eAkz["eintraege"]}

    # Aufruf Projektmasterdatei einzulesen und auszuwerten
    alle_ergebnisse = projektmaster_einlesen(input_dir, data)

    # Ergebnisse in die Statusdatei schreiben
    statusdatei = status_excel_erstellen(data, alle_ergebnisse)

    adressen_auswerten(statusdatei, input_dir, data)

    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Ende der Verarbeitung")

def projektmaster_einlesen(inputdir, data):

    pm_datei = Path(inputdir + "ProjektMasterDatei_2026_05_05.xlsx")

    alle_ergebnisse = []

    #Landkreis ADK einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe ADK", header=data['ADK_Header'])
    #print(df.head())
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten ADK eingelesen")

    # Landkreis LBC _neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe LBC _neu", header=data['LBC_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten LBC eingelesen")

    # Landkreis FDS einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe FDS_neu", header=data['FDS_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten FDS eingelesen")

    # Landkreis LRT_neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe LRT_neu", header=data['LRT_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten LRT eingelesen")

    #Landkreis SIG einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe SIG", header=data['SIG_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten SIG eingelesen")

    #Landkreis ZAK_neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe ZAK_neu", header=data['ZAK_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten ZAK eingelesen")

    return alle_ergebnisse


def projektmaster_details(df, alle_ergebnisse):
    df.columns = df.columns.str.strip()

    # Name der ersten Spalte ermitteln
    erste_spalte = df.columns[0]
    sechste_spalte = df.columns[5]
    # Hier sammeln wir die Treffer

    for index, row in df.iterrows():

        wert_a = row[erste_spalte]
        wert_gemeinde = str(row[sechste_spalte])

        # Zeilen die mit "Alle" in der Spalte Gemeinde beginnen überspringen
        # Zeilen werden doch benötigt, da die Controlling-Liste sie enthält
        if wert_gemeinde.startswith("alle"):
            continue

        # Leere Zeilen überspringen
        if pd.isna(wert_a) or str(wert_a).strip() == "":
            continue

        # In String umwandeln
        wert_a = str(wert_a).strip()

        # Abbruchbedingung
        if wert_a.lower() == "möglicher status:":
            #print("Ende erreicht")
            break

        # Relevante Zeile speichern
        alle_ergebnisse.append({
            "eAktenzeichen": row["eAZ"],
            "Landkreis": row["LK-Kürzel"],
            "Gemeinde": row["Gemeinde"],
            "Gemarkung": row["Gemarkung"],
            "Ausschreibung": row["Ausschreibung"],
            "Cluster": row["Cluster"],
            "Los": row["Los"],
            "Bauclustercode Kfm.": row["Bauclustercode Kfm."],
            "Planung Beginn": row.iloc[19],
            "Planung Ende": row.iloc[20],
            "Bauphase Beginn": row.iloc[24],
            "Bauphase Ende": row.iloc[25],

        })

    return alle_ergebnisse


def status_excel_erstellen(data, alle_ergebnisse):

    gesamt_df = pd.DataFrame(alle_ergebnisse)

    template = home + "/" + data['output_dir'] + data['status_template']  #Status_Template
    sheet = "Master"

    datum = datetime.today().strftime("%Y%m%d")
    ausgabe = home + "/" + data['output_dir'] + f"AP24_Master_Status_{datum}.xlsx"

    # Zielspalten in Excel (0-basiert!)
    column_mapping = {
        "eAktenzeichen": 2,
        "Landkreis": 3,
        "Gemeinde": 4,
        "Gemarkung": 5,
        "Ausschreibung": 7,
        "Cluster": 8,
        "Los": 9,
        "Bauclustercode Kfm.": 10,
        "Planung Beginn": 11,
        "Planung Ende": 12,
        "Bauphase Beginn": 13,
        "Bauphase Ende": 14,

    }

    # Existierende Datei laden (Formatierung & Formeln bleiben erhalten)
    wb = load_workbook(template)
    ws = wb[sheet]  # ggf. Sheet-Name anpassen

    START_ROW = 5  # Zeile, ab der Daten geschrieben werden (1 = erste Zeile)

    for df_col, excel_col in column_mapping.items():
        for row_idx, value in enumerate(gesamt_df[df_col], start=START_ROW):
            cell = ws.cell(row=row_idx, column=excel_col, value=value)
            # Datumspalten formatieren
            if df_col in ["Planung Beginn", "Planung Ende", "Bauphase Beginn", "Bauphase Ende"]:
                cell.number_format = "DD.MM.YYYY"

    end_row = START_ROW + len(gesamt_df) - 1

    for row_idx in range(START_ROW, end_row + 1):
        ws.cell(row=row_idx, column=6, value=f'=VLOOKUP(B{row_idx},Daten!Z$1:AA$57,2,)') # Formel muss in Englisch angegeben werden

    wb.save(ausgabe)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Basisdaten in Statusdatei befüllt")

    return ausgabe


def adressen_auswerten(statusdatei, input_dir, data):
    STATUS_DATEI = statusdatei # gefüllte "AP24_Master_Status..."
    QUELL_DATEI = input_dir + "Masterliste.xlsx" # Aktuell nur die ADK Datei, Muss erweitert werden
    STATUS_SHEET = "Master"
    QUELL_SHEET = "Adressliste_zur Bearbeitung_int" #Exceklsheet, welches ausgewertet wird
    START_ROW = 5

    ADRESSEN_BASIS_COL = 15
    ADRESSEN_HINZUNAHME_COL = 16
    ADRESSEN_ENTNAHME_COL = 17
    ADRESSEN_FM_IN_BEARBEITUNG = 18
    ADRESSEN_FM_ERLEDIGT = 20
    ADRESSEN_AKTUELL = 21

    GU = 11 # Spalte K - Hier kommt der GÜ/GU rein
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Start Auswertung der Adressdaten je Landkreis")

    #print(df.head())

    # Dynamischer Suchwert kommt aus Spalte J der Basis → sucht in dieser Quellspalte
    SUCH_SPALTE_QUELLE = "bauabschnitt_code"  # Spaltenname in der Quelldatei Spalte CH

    def adressen_basis_treffer(suchwert, df):
        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "bauabschnitt_code": suchwert,  # Spaltenname in Quelldatei : gesuchter Wert
            # weitere Kriterien einfach ergänzen...
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        #if suchwert == "ADK_S_02-":
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]
        #gefiltert = df[df[SUCH_SPALTE_QUELLE].astype(str).str.startswith(str(suchwert))]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte] != "")]
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]

        #gefiltert = df[df[SUCH_SPALTE_QUELLE].astype(str).str.startswith(str(suchwert))]
        #anzahl = len(gefiltert)

        return len(gefiltert)

    def adressen_hinzunahme_gu(suchwert, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "Adressänderung": "Hinzunahme",  # Spaltenname in Quelldatei : gesuchter Wert Förderverfahrenszuordnung
            # weitere Kriterien einfach ergänzen...
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte] != "")]
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_entnahme_gu(suchwert, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "Adressänderung": "Entnahme",  # Spaltenname in Quelldatei : gesuchter Wert Förderverfahrenszuordnung
            # weitere Kriterien einfach ergänzen...
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]

        for spalte, wert in fixe_kriterien.items():
            gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_fm_in_bearbeitung(suchwert, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "status_wechsel_pt": ["in Bearbeitung", "Angezeigt bei PT", "Bescheid liegt vor"],  # Spaltenname in Quelldatei : gesuchter Wert Förderverfahrenszuordnung
            # weitere Kriterien einfach ergänzen...
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte].astype(str).str.strip() != "")]
            elif isinstance(wert, list):
                gefiltert = gefiltert[gefiltert[spalte].isin(wert)]  # ← mehrere Werte
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_fm_erledigt(suchwert, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "status_wechsel_pt": ["Aus Antrag entfernen", "In Antrag hinzunehmen"],
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte].astype(str).str.strip() != "")]
            elif isinstance(wert, list):
                gefiltert = gefiltert[gefiltert[spalte].isin(wert)]  # ← mehrere Werte
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_fm_aus_antrag_entfernen(suchwert, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "status_wechsel_pt": "Aus Antrag entfernen",
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]

        for spalte, wert in fixe_kriterien.items():
            gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_fm_in_antrag_hinzunehmen(suchwert, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "status_wechsel_pt": "In Antrag hinzunehmen",
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]

        for spalte, wert in fixe_kriterien.items():
            gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    # Basis-Excel laden
    wb = load_workbook(STATUS_DATEI)
    ws = wb[STATUS_SHEET]

    address_master = [home + "/" + data["input_dir"] + data["adressablage"] + data["ADK_ADRESS_MASTER"],
                      home + "/" + data["input_dir"] + data["adressablage"] + data["LBC_ADRESS_MASTER"],
                      home + "/" + data["input_dir"] + data["adressablage"] + data["FDS_ADRESS_MASTER"],
                      home + "/" + data["input_dir"] + data["adressablage"] + data["LRT_ADRESS_MASTER"],
                      home + "/" + data["input_dir"] + data["adressablage"] + data["SIG_ADRESS_MASTER"],
                      home + "/" + data["input_dir"] + data["adressablage"] + data["ZAK_ADRESS_MASTER"]
                      ]
    #address_master = [home + "/" + data["input_dir"] + data["adressablage"] + data["ADK_ADRESS_MASTER"]]
    #                  home + "/" + data["input_dir"] + data["adressablage"] + data["LBC_ADRESS_MASTER"]]

    for quell_datei in address_master:
        datum = datetime.today().strftime("%Y%m%d %H%M%S")
        print(datum + ": Adressdatei: " + quell_datei + " gestartet")

        df = pd.read_excel(quell_datei, sheet_name=QUELL_SHEET, header=11)
        #print(df.head())

        for row_idx in range(START_ROW, ws.max_row + 1):
            suchwert = ws.cell(row=row_idx, column=10).value  # Spalte J = 10

            if suchwert is None or str(suchwert).strip() == "":
                continue
            #print(suchwert)
            anzahl_basis_value = adressen_basis_treffer(suchwert, df)
            anzahl_hinzunahme_value = adressen_hinzunahme_gu(suchwert, df)
            anzahl_entnahme_value = adressen_entnahme_gu(suchwert, df)
            anzahl_fm_in_bearbeitung_value = adressen_fm_in_bearbeitung(suchwert, df)
            anzahl_fm_erledigt_value = adressen_fm_erledigt(suchwert, df)
            anzahl_fm_aus_antrag_entfernen = adressen_fm_aus_antrag_entfernen(suchwert, df)
            anzahl_fm_in_antrag_hinzunehmen = adressen_fm_in_antrag_hinzunehmen(suchwert, df)
            anzahl_aktuell_value = anzahl_basis_value - anzahl_fm_aus_antrag_entfernen + anzahl_fm_in_antrag_hinzunehmen
            #gu_wert = ws.cell(row=row_idx, column=GU).value
            #print(f"{anzahl_aktuell_value} und {anzahl_basis_value} und {anzahl_fm_aus_antrag_entfernen} und {anzahl_fm_in_antrag_hinzunehmen}")
            if anzahl_basis_value != 0:
                ws.cell(row=row_idx, column=ADRESSEN_BASIS_COL, value=anzahl_basis_value)
            if anzahl_hinzunahme_value != 0:
                ws.cell(row=row_idx, column=ADRESSEN_HINZUNAHME_COL, value=anzahl_hinzunahme_value)
            if anzahl_entnahme_value != 0:
                ws.cell(row=row_idx, column=ADRESSEN_ENTNAHME_COL, value=anzahl_entnahme_value)
            if anzahl_fm_in_bearbeitung_value != 0:
                ws.cell(row=row_idx, column=ADRESSEN_FM_IN_BEARBEITUNG, value=anzahl_fm_in_bearbeitung_value)
            if anzahl_fm_erledigt_value != 0:
                ws.cell(row=row_idx, column=ADRESSEN_FM_ERLEDIGT, value=anzahl_fm_erledigt_value)
            if anzahl_aktuell_value != 0:
                ws.cell(row=row_idx, column=ADRESSEN_AKTUELL, value=anzahl_aktuell_value)


    wb.save(STATUS_DATEI)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Adressmaster-Daten verarbeitet")


if __name__ == '__main__':
    main()