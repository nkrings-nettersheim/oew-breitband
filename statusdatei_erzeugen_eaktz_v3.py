import json
import os
import sys

import pandas as pd
import warnings

from openpyxl import load_workbook
from openpyxl.styles import numbers
from pathlib import Path
from datetime import datetime

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

home = os.path.expanduser('~')

def main(prod):

    datum = datetime.today().strftime("%Y%m%d %H%M%S")

    print(datum + ": Start der Verarbeitung")

    my_dir = Path(__file__).parent

    # Konfiguration einlesen
    f = open(my_dir / 'config_v2.json')
    data = json.load(f)
    data['Prod']=prod
    f.close()

    # Aufruf Projektmasterdatei einzulesen und auszuwerten
    #alle_ergebnisse = projektmaster_einlesen(data)

    # Ergebnisse in die Statusdatei schreiben
    #statusdatei = status_excel_erstellen(data, alle_ergebnisse)

    statusdatei = home + "/" + data['output_dir'] + f"AP24_Master_Status_V2.xlsx"

    adressen_auswerten(data, statusdatei)

    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Ende der Verarbeitung")

def projektmaster_einlesen(data):

    if data["Prod"] == "j":
        pm_datei = Path( home + "/" + data["oew_ablage_pmd"] + data['pmd_datei'])
    else:
        pm_datei = Path( home + "/" + data["input_dir"] + data['pmd_datei'])

    alle_ergebnisse = []
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PMD-Datei " + str(pm_datei) + " einlesen")

    #Landkreis ADK einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe ADK", header=data['ADK_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten ADK eingelesen")

    # Landkreis LBC _neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe LBC _neu", header=data['LBC_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten LBC eingelesen")

    # Landkreis FDS einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe FDS", header=data['FDS_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten FDS eingelesen")

    # Landkreis LRT_neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe LRT", header=data['LRT_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten LRT eingelesen")

    #Landkreis SIG einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe SIG", header=data['SIG_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten SIG eingelesen")

    #Landkreis ZAK_neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe ZAK", header=data['ZAK_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten ZAK eingelesen")

    return alle_ergebnisse


def projektmaster_details(df, alle_ergebnisse):
    df.columns = df.columns.str.strip()

    # Name der ersten Spalte ermitteln
    erste_spalte = df.columns[0]
    sechste_spalte = df.columns[5]

    # Hier sammeln wir die Treffer
    for index, row in df.iterrows():

        wert_a = row[erste_spalte]
        wert_gemeinde = str(row[sechste_spalte])
        if wert_gemeinde == "8425002":
            print("Hallo")

        # Zeilen die mit "Alle" in der Spalte Gemeinde beginnen überspringen
        if wert_gemeinde.startswith("alle"):
            continue

        # Leere Zeilen überspringen
        if pd.isna(wert_a) or str(wert_a).strip() == "":
            continue

        if pd.isna(wert_gemeinde) or str(wert_gemeinde).strip() == "":
            continue

        # Zeilen in denen etwas anderes steht als eine eAZ und nicht leer sind entfernen
        if pd.notna(row["eAZ"]) and not str(row["eAZ"]).startswith("832."):
            continue

        # In String umwandeln
        wert_a = str(wert_a).strip()

        # Abbruchbedingung
        if wert_a.lower() == "möglicher status:":
            break

        # Relevante Zeile speichern
        alle_ergebnisse.append({
            "eAktenzeichen": row["eAZ"],
            "Landkreis": row["LK-Kürzel"],
            "Gemeinde": row["Gemeinde"],
            "Gemarkung": row["Gemarkung"],
            "Ausschreibung": row["Ausschreibung"],
            "Cluster": row["Cluster"],
            "Los": row["Los"],
            "Bauclustercode Kfm.": row["Bauclustercode Kfm."],
            "Planung Beginn": row.iloc[19],
            "Planung Ende": row.iloc[20],
            "Bauphase Beginn": row.iloc[24],
            "Bauphase Ende": row.iloc[25],
        })

    return alle_ergebnisse


def status_excel_erstellen(data, alle_ergebnisse):
    # Umwandlung der Daten in einen Dataframe
    df = pd.DataFrame(alle_ergebnisse)

    # Duplicate entfernen die aufgrund mehrerer Baucluster entstehen
    gesamt_df = df.drop_duplicates(subset=['eAktenzeichen', 'Gemeinde', 'Gemarkung'])

    # Sortieren der Daten nach Landkreis, Ausschreibung, Gemeinde und Gemarkung
    gesamt_df = gesamt_df.sort_values(by=['Landkreis','Ausschreibung' ,'Gemeinde', 'Gemarkung'])

    # Pfad und Dateiname definieren für die Ausgabe
    template = home + "/" + data['output_dir'] + data['status_template']  #Status_Template
    sheet = "Master"
    datum = datetime.today().strftime("%Y%m%d")
    ausgabe = home + "/" + data['output_dir'] + f"AP24_Master_Status_{datum}_V2.xlsx"

    # Zielspalten in Excel (0-basiert!)
    column_mapping = {
        "eAktenzeichen": 1,
        "Landkreis": 2,
        "Gemeinde": 3,
        "Gemarkung": 4,
        "Ausschreibung": 6,
        "Cluster": 7,
        "Los": 8,
        "Bauclustercode Kfm.": 9,
        "Planung Beginn": 10,
        "Planung Ende": 11,
        "Bauphase Beginn": 12,
        "Bauphase Ende": 13,
    }

    # Existierende Datei laden (Formatierung & Formeln bleiben erhalten)
    wb = load_workbook(template)
    ws = wb[sheet]  # ggf. Sheet-Name anpassen

    START_ROW = 5  # Zeile, ab der Daten geschrieben werden (1 = erste Zeile)

    # Datumssplaten formatieren
    for df_col, excel_col in column_mapping.items():
        for row_idx, value in enumerate(gesamt_df[df_col], start=START_ROW):
            cell = ws.cell(row=row_idx, column=excel_col, value=value)
            # Datumspalten formatieren
            if df_col in ["Planung Beginn", "Planung Ende", "Bauphase Beginn", "Bauphase Ende"]:
                cell.number_format = "DD.MM.YYYY"

    END_ROW = START_ROW + len(gesamt_df) - 1

    # In Spalte 6 einen SVERWEIS einbauen
    # Dieser schaut in einer Referenztabelle nach, da die Info in dem Block der Daten der PMD nicht enthalten ist
    for row_idx in range(START_ROW, END_ROW + 1):
        ws.cell(row=row_idx, column=5, value=f'=VLOOKUP(A{row_idx},Daten!Z$1:AA$57,2,)') # Formel muss in Englisch angegeben werden

    wb.save(ausgabe)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Basisdaten in Statusdatei befüllt")

    return ausgabe


def adressen_auswerten(data, statusdatei):
    STATUS_DATEI = statusdatei # gefüllte "AP24_Master_Status..."
    STATUS_SHEET = "Master"
    DETAIL_SHEET = "Details_Aenderungen"
    QUELL_SHEET = data["Quellsheet_Adressmaster"] #Excelsheet, welches ausgewertet wird
    START_ROW = 5

    ADRESSEN_GESAMT_COL = 13
    ADRESSEN_BASIS_COL = 14
    ADRESSEN_ENTNAHME = 15
    ADRESSEN_HINZUNAHME = 16
    ADRESSEN_IN_BEARBEITUNG_FM = 17
    ADRESSEN_IN_BEARBEITUNG_PT = 18
    ADRESSEN_ANGENOMMEN = 19
    ADRESSEN_ABGELEHNT = 20
    ADRESSEN_AKTUELL = 21
    ADRESSEN_Y_COL = 22
    ADRESSEN_X_COL = 23


    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Start Auswertung der Adressdaten je Landkreis")

    # Erstes Kriterium, nachdem die Masterdatei gefiltert wird
    SUCH_SPALTE_QUELLE = "ortsname"

    def adressen_gesamt(gemeinde, gemarkung, eaktz, bc_code, df):

        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "hgf_aktenzeichnen_quelle" : eaktz,
            "bauabschnitt_code": bc_code,
        }

        # Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück.
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]
        #if gemarkung == "Ehlenbogen":
        #    print("Hallo")
        for spalte, wert in fixe_kriterien.items():
            if wert == "":
                gefiltert = gefiltert[gefiltert[spalte].isna()]
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]
        wert_hgf = len(gefiltert)

        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "dgf_aktenzeichnen_quelle": eaktz,
            "bauabschnitt_code": bc_code,
        }

        # Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück.
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]
        # if gemarkung == "Ehlenbogen":
        #    print("Hallo")
        for spalte, wert in fixe_kriterien.items():
            if wert == "":
                gefiltert = gefiltert[gefiltert[spalte].isna()]
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]
        wert_dgf = len(gefiltert)

        basis_wert = wert_hgf + wert_dgf
        return basis_wert

    def adressen_basis_treffer(gemeinde, gemarkung, eaktz, bc_code, df):
        # hgf Abschnitt
        # weitere Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        #if gemarkung == "Aach":
        #    print('Hallo')

        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "hgf_aktenzeichnen_quelle" : eaktz,
            "bauabschnitt_code": bc_code,
            "hgf_untervers": 1,
        }

        # Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück.
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte] != "")]
            elif wert == "LEER":
                gefiltert = gefiltert[gefiltert[spalte].isna() | (gefiltert[spalte].astype(str).str.strip() == "")]
            elif wert == "":
                gefiltert = gefiltert[gefiltert[spalte].isna()]
            else:
                if spalte == "hgf_untervers":
                    gefiltert[spalte] = pd.to_numeric(gefiltert[spalte], errors='coerce')
                gefiltert = gefiltert[gefiltert[spalte] == wert]
        wert_hgf = len(gefiltert)

        #if gemarkung == "Berghülen":
        #    print('Hallo')

        # dgf Abschnitt
        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "dgf_aktenzeichnen_quelle": eaktz,
            "bauabschnitt_code": bc_code,
            "dgf_untervers": 1,
        }

        # Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück.
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte] != "")]
            elif wert == "":
                gefiltert = gefiltert[gefiltert[spalte].isna()]
            else:
                if spalte == "dgf_untervers":
                    gefiltert[spalte] = pd.to_numeric(gefiltert[spalte], errors='coerce')
                gefiltert = gefiltert[gefiltert[spalte] == wert]
        wert_dgf = len(gefiltert)

        basis_wert = wert_hgf + wert_dgf
        return basis_wert

    def adressen_detail(gemeinde, gemarkung, eaktz, bc_code, df):

        STATUS_WECHSEL_PT_NEU = ['In Bearbeitung FM', 'In Bearbeitung PT','Angenommen', 'Abgelehnt']
        AENDERUNG_PT = ['Entnahme -> EWA',
                        'Entnahme -> nicht förderfähig',
                        'Entnahme -> Mini-MEV',
                        'Entnahme -> HFC-Nachschärfung',
                        'Entnahme -> sonstiges',
                        'Hinzunahme -> Mini-MEV',
                        'Hinzunahme -> Neue Adresse',
                        'Hinzunahme -> HFC-Nachschärfung',
                        'Hinzunahme -> sonstiges',
                        'Andere Gründe']

        result = []

        for status_wechsel_pt_neu in STATUS_WECHSEL_PT_NEU:
            for aenderung_pt in AENDERUNG_PT:
                # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
                fixe_kriterien = {
                    "gemark_nam": gemarkung,
                    "foerderazbnd": eaktz,
                    "bauabschnitt_code": bc_code,
                    "status_wechsel_pt_neu": status_wechsel_pt_neu,
                    "aenderung_pt": aenderung_pt,
                }

                """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
                gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

                for spalte, wert in fixe_kriterien.items():
                    if wert == "NICHT_LEER":
                        gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte].astype(str).str.strip() != "")]
                    elif wert == "LEER":
                        gefiltert = gefiltert[gefiltert[spalte].isna() | (gefiltert[spalte].astype(str).str.strip() == "")]
                    elif wert == "":
                        gefiltert = gefiltert[gefiltert[spalte].isna()]
                    else:
                        gefiltert = gefiltert[gefiltert[spalte] == wert]
                result.append((status_wechsel_pt_neu, aenderung_pt, len(gefiltert)))

        return result

    def adressen_aktuell(gemeinde, gemarkung, eaktz, bc_code, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "status_wechsel_pt": "LEER",
            "foerderazbnd": eaktz,
            "bauabschnitt_code": bc_code,
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte].astype(str).str.strip() != "")]
            elif wert == "LEER":
                gefiltert = gefiltert[gefiltert[spalte].isna() | (gefiltert[spalte].astype(str).str.strip() == "")]
            elif wert == "":
                gefiltert = gefiltert[gefiltert[spalte].isna()]
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def details_schreiben(ws_details, row_idx, eaktz, gemeinde, gemarkung, anzahl_details):
        # Startspalte für die Detailswerte
        x=4
        ws_details.cell(row=row_idx, column=1, value=eaktz)
        ws_details.cell(row=row_idx, column=2, value=gemeinde)
        ws_details.cell(row=row_idx, column=3, value=gemarkung)
        for status_wechsel_pt_neu, aenderung_pt, anzahl in anzahl_details:
            ws_details.cell(row=row_idx, column=x, value=anzahl)
            x= x+1

    # Basis-Excel laden
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Statusdatei: " + STATUS_DATEI + " wird geladen")
    wb = load_workbook(STATUS_DATEI)
    ws_master = wb[STATUS_SHEET]
    ws_details = wb[DETAIL_SHEET]

    if data['Prod'] == "j":
        address_master = [home + "/" + data["oew_ablage_adressen"] + data["ADK_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["BC_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["FDS_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["REU_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["SIG_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["ZAK_ADRESS_MASTER"]
                          ]
    else:
        address_master = [home + "/" + data["input_dir"] + data["ADRESS_MASTER_TEST"]]
        # address_master = [home + "/" + data["oew_ablage_adressen"] + data["LBC_ADRESS_MASTER"]]
        # address_master = [home + "/" + data["oew_ablage_adressen"] + data["FDS_ADRESS_MASTER"]]

    for quell_datei in address_master:
        datum = datetime.today().strftime("%Y%m%d %H%M%S")
        print(datum + ": Adressdatei: " + quell_datei + " gestartet")

        # Ermitteln welcher Landkreis gerade ausgewertet wird um die Suche darauf einzuschränken
        if quell_datei == home + "/" + data["oew_ablage_adressen"] + data["ADK_ADRESS_MASTER"]:
            lk = "ADK"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["BC_ADRESS_MASTER"]:
            lk = "BC"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["FDS_ADRESS_MASTER"]:
            lk = "FDS"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["REU_ADRESS_MASTER"]:
            lk = "REU"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["SIG_ADRESS_MASTER"]:
            lk = "SIG"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["ZAK_ADRESS_MASTER"]:
            lk = "ZAK"
        else:
            lk = "FDS"

        df = pd.read_excel(quell_datei, sheet_name=QUELL_SHEET, header=14)

        for row_idx in range(START_ROW, ws_master.max_row + 1):
            gemeinde = (ws_master.cell(row=row_idx, column=3).value or "").rstrip()
            gemarkung = (ws_master.cell(row=row_idx, column=4).value or "").rstrip()
            eaktz = (ws_master.cell(row=row_idx, column=1).value or "").rstrip()
            landkreis = (ws_master.cell(row=row_idx, column=2).value or "").rstrip()
            bc_code = (ws_master.cell(row=row_idx, column=8).value or "").rstrip()

            if landkreis and landkreis.startswith(lk):
                if gemeinde is None or str(gemeinde).strip() == "":
                    continue

                datum = datetime.today().strftime("%Y%m%d %H%M%S")
                print(datum + ": Auswertung Gemeinde/Gemarkung: /" + str(gemeinde) + "/" + str(gemarkung) +"/")

                anzahl_gesamt = adressen_gesamt(gemeinde, gemarkung, eaktz, bc_code, df)
                anzahl_basis_value = adressen_basis_treffer(gemeinde, gemarkung, eaktz, bc_code, df)
                anzahl_details = adressen_detail(gemeinde, gemarkung, eaktz, bc_code, df)
                anzahl_aktuell_value = adressen_aktuell(gemeinde, gemarkung, eaktz, bc_code, df)


                details_schreiben(ws_details, row_idx, eaktz, gemeinde, gemarkung, anzahl_details)
                anzahl_entnahme = sum(anzahl for status_wechsel_pt_neu, aenderung_pt, anzahl in anzahl_details
                    if aenderung_pt in ('Entnahme -> EWA',
                        'Entnahme -> nicht förderfähig',
                        'Entnahme -> Mini-MEV',
                        'Entnahme -> HFC-Nachschärfung',
                        'Entnahme -> sonstiges'
                        )
                )
                anzahl_hinzunahme = sum(anzahl for status_wechsel_pt_neu, aenderung_pt, anzahl in anzahl_details
                    if aenderung_pt in ('Hinzunahme -> Mini-MEV',
                        'Hinzunahme -> Neue Adresse',
                        'Hinzunahme -> HFC-Nachschärfung',
                        'Hinzunahme -> sonstiges'
                        )
                )
                anzahl_in_bearbeitung_fm = sum(anzahl for status_wechsel_pt_neu, aenderung_pt, anzahl in anzahl_details
                                                if status_wechsel_pt_neu == "In Bearbeitung FM")
                anzahl_in_bearbeitung_pt = sum(anzahl for status_wechsel_pt_neu, aenderung_pt, anzahl in anzahl_details
                                               if status_wechsel_pt_neu == "In Bearbeitung PT")
                anzahl_angenommen = sum(anzahl for status_wechsel_pt_neu, aenderung_pt, anzahl in anzahl_details
                                               if status_wechsel_pt_neu == "Angenommen")
                anzahl_abgelehnt = sum(anzahl for status_wechsel_pt_neu, aenderung_pt, anzahl in anzahl_details
                                               if status_wechsel_pt_neu == "Abgelehnt")

                if anzahl_gesamt != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_GESAMT_COL, value=anzahl_gesamt)

                if anzahl_basis_value != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_BASIS_COL, value=anzahl_basis_value)

                if anzahl_entnahme != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_ENTNAHME, value=anzahl_entnahme)

                if anzahl_hinzunahme != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_HINZUNAHME, value=anzahl_hinzunahme)

                if anzahl_in_bearbeitung_fm != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG_FM, value=anzahl_in_bearbeitung_fm)

                if anzahl_in_bearbeitung_pt != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG_PT, value=anzahl_in_bearbeitung_pt)

                if anzahl_angenommen != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_ANGENOMMEN, value=anzahl_angenommen)

                if anzahl_abgelehnt != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_ABGELEHNT, value=anzahl_abgelehnt)

                if anzahl_aktuell_value != 0:
                    ws_master.cell(row=row_idx, column=ADRESSEN_AKTUELL, value=anzahl_aktuell_value)

                # Logik für Ermittlung Baubeginn
                ws_master.cell(row=row_idx, column=ADRESSEN_Y_COL, value=2)

                anteil_bearbeitung = (anzahl_in_bearbeitung_fm + anzahl_in_bearbeitung_pt) / anzahl_basis_value
                ws_master.cell(row=row_idx, column=ADRESSEN_X_COL, value=anteil_bearbeitung)


    for row_idx in range(START_ROW, ws_master.max_row + 1):
        if ws_master.cell(row=row_idx, column=ADRESSEN_GESAMT_COL).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_GESAMT_COL, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_BASIS_COL).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_BASIS_COL, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_ENTNAHME).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_ENTNAHME, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_HINZUNAHME).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_HINZUNAHME, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG_FM).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG_FM, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG_PT).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG_PT, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_ANGENOMMEN).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_ANGENOMMEN, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_ABGELEHNT).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_ABGELEHNT, value=0)
        if ws_master.cell(row=row_idx, column=ADRESSEN_AKTUELL).value is None:
            ws_master.cell(row=row_idx, column=ADRESSEN_AKTUELL, value=0)

    ausgabe = home + "/" + data['output_dir'] + f"AP24_Master_Status_{datum}_V3.xlsx"
    wb.save(ausgabe)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Adressmaster-Daten verarbeitet")


if __name__ == '__main__':
    prod = sys.argv[1]
    main(prod)