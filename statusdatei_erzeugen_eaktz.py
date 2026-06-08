import json
import os
import pandas as pd
import warnings

from openpyxl import load_workbook
from openpyxl.styles import numbers
from pathlib import Path
from datetime import datetime

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

home = os.path.expanduser('~')


def main():
    #home = os.path.expanduser('~')

    datum = datetime.today().strftime("%Y%m%d %H%M%S")

    print(datum + ": Start der Verarbeitung")

    my_dir = Path(__file__).parent

    # Konfiguration einlesen
    f = open(my_dir / 'config.json')
    data = json.load(f)
    input_dir = home + "/" + data['input_dir']
    output_dir = home + "/" + data['output_dir']
    f.close()

    # Liste der eAkz einlesen
    #f = open(my_dir / 'eAktenzeichen.json')
    #eAkz = json.load(f)
    #f.close()
    #eAkz_dictionary = {eintrag["aktenzeichen"]: eintrag for eintrag in eAkz["eintraege"]}

    # Aufruf Projektmasterdatei einzulesen und auszuwerten
    alle_ergebnisse = projektmaster_einlesen(data)

    # Ergebnisse in die Statusdatei schreiben
    statusdatei = status_excel_erstellen(data, alle_ergebnisse)

    #adressen_auswerten(data, statusdatei)

    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Ende der Verarbeitung")


def projektmaster_einlesen(data):
    pm_datei = Path( home + "/" + data["oew_ablage_pmd"] + data['pmd_datei'])

    alle_ergebnisse = []

    #Landkreis ADK einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe ADK", header=data['ADK_Header'])
    #print(df.head())
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten ADK eingelesen")

    # Landkreis LBC _neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe LBC _neu", header=data['LBC_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten LBC eingelesen")

    # Landkreis FDS einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe FDS", header=data['FDS_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten FDS eingelesen")

    # Landkreis LRT_neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe LRT", header=data['LRT_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten LRT eingelesen")

    #Landkreis SIG einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe SIG", header=data['SIG_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten SIG eingelesen")

    #Landkreis ZAK_neu einlesen
    df = pd.read_excel(pm_datei, sheet_name="Planungs und Bauabrufe ZAK", header=data['ZAK_Header'])
    alle_ergebnisse = projektmaster_details(df, alle_ergebnisse)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": PM-Daten ZAK eingelesen")

    return alle_ergebnisse


def projektmaster_details(df, alle_ergebnisse):
    df.columns = df.columns.str.strip()

    # Name der ersten Spalte ermitteln
    erste_spalte = df.columns[0]
    sechste_spalte = df.columns[5]
    # Hier sammeln wir die Treffer

    for index, row in df.iterrows():

        wert_a = row[erste_spalte]
        wert_gemeinde = str(row[sechste_spalte])

        # Zeilen die mit "Alle" in der Spalte Gemeinde beginnen überspringen
        # Zeilen werden doch benötigt, da die Controlling-Liste sie enthält
        if wert_gemeinde.startswith("alle"):
            continue

        # Leere Zeilen überspringen
        if pd.isna(wert_a) or str(wert_a).strip() == "":
            continue

        if pd.notna(row["eAZ"]) and not str(row["eAZ"]).startswith("832."):
            continue

        # In String umwandeln
        wert_a = str(wert_a).strip()

        # Abbruchbedingung
        if wert_a.lower() == "möglicher status:":
            #print("Ende erreicht")
            break

        # Relevante Zeile speichern
        alle_ergebnisse.append({
            "eAktenzeichen": row["eAZ"],
            "Landkreis": row["LK-Kürzel"],
            "Gemeinde": row["Gemeinde"],
            "Gemarkung": row["Gemarkung"],
            "Ausschreibung": row["Ausschreibung"],
            "Cluster": row["Cluster"],
            "Los": row["Los"],
            "Bauclustercode Kfm.": row["Bauclustercode Kfm."],
            "Planung Beginn": row.iloc[19],
            "Planung Ende": row.iloc[20],
            "Bauphase Beginn": row.iloc[24],
            "Bauphase Ende": row.iloc[25],

        })

    return alle_ergebnisse


def status_excel_erstellen(data, alle_ergebnisse):
    # Umwandlung der Daten in einen Dataframe
    df = pd.DataFrame(alle_ergebnisse)

    # Duplicate entfernen die aufgrund mehrerer Baucluster entstehen
    gesamt_df = df.drop_duplicates(subset=['eAktenzeichen', 'Gemeinde', 'Gemarkung'])

    # Sortieren der DAten nach Landkreis, Ausschreibung, Gemeinde und Gemarkung
    gesamt_df = gesamt_df.sort_values(by=['Landkreis', 'Ausschreibung', 'Gemeinde', 'Gemarkung'])

    template = home + "/" + data['output_dir'] + data['status_template']  #Status_Template
    sheet = "Master"

    datum = datetime.today().strftime("%Y%m%d")
    ausgabe = home + "/" + data['output_dir'] + f"AP24_Master_Status_{datum}.xlsx"

    # Zielspalten in Excel (0-basiert!)
    column_mapping = {
        "eAktenzeichen": 2,
        "Landkreis": 3,
        "Gemeinde": 4,
        "Gemarkung": 5,
        "Ausschreibung": 7,
        "Cluster": 8,
        "Los": 9,
        "Bauclustercode Kfm.": 10,
        "Planung Beginn": 11,
        "Planung Ende": 12,
        "Bauphase Beginn": 13,
        "Bauphase Ende": 14,

    }

    # Existierende Datei laden (Formatierung & Formeln bleiben erhalten)
    wb = load_workbook(template)
    ws = wb[sheet]  # ggf. Sheet-Name anpassen

    START_ROW = 5  # Zeile, ab der Daten geschrieben werden (1 = erste Zeile)

    for df_col, excel_col in column_mapping.items():
        for row_idx, value in enumerate(gesamt_df[df_col], start=START_ROW):
            cell = ws.cell(row=row_idx, column=excel_col, value=value)
            # Datumspalten formatieren
            if df_col in ["Planung Beginn", "Planung Ende", "Bauphase Beginn", "Bauphase Ende"]:
                cell.number_format = "DD.MM.YYYY"

    end_row = START_ROW + len(gesamt_df) - 1

    for row_idx in range(START_ROW, end_row + 1):
        ws.cell(row=row_idx, column=6,
                value=f'=VLOOKUP(B{row_idx},Daten!Z$1:AA$57,2,)')  # Formel muss in Englisch angegeben werden

    wb.save(ausgabe)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Basisdaten in Statusdatei befüllt")

    return ausgabe


def adressen_auswerten(data, statusdatei):
    STATUS_DATEI = statusdatei  # gefüllte "AP24_Master_Status..."
    STATUS_SHEET = "Master"
    QUELL_SHEET = "Adressliste_zur Bearbeitung_int"  #Excelsheet, welches ausgewertet wird
    START_ROW = 5

    ADRESSEN_BASIS_COL = 15
    ADRESSEN_HINZUNAHME_COL = 16
    ADRESSEN_ENTFERNEN_COL = 17
    ADRESSEN_IN_BEARBEITUNG = 18
    ADRESSEN_AKTUELL = 19

    GU = 11  # Spalte K - Hier kommt der GÜ/GU rein
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Start Auswertung der Adressdaten je Landkreis")

    # Dynamischer Suchwert kommt aus Spalte J der Basis → sucht in dieser Quellspalte
    SUCH_SPALTE_QUELLE = "ortsname"  # Spaltenname in der Quelldatei Spalte CH

    def adressen_basis_treffer(gemeinde, gemarkung, eaktz, df):
        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "hgf_aktenzeichnen_quelle": eaktz,
            "hgf_untervers": 1,
        }
        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte] != "")]
            else:
                if spalte == "hgf_untervers":
                    gefiltert[spalte] = pd.to_numeric(gefiltert[spalte], errors='coerce')
                gefiltert = gefiltert[gefiltert[spalte] == wert]
        wert_hgf = len(gefiltert)

        #if gemarkung == "Dietenheim ":
        #    print('Hallo')

        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "dgf_aktenzeichnen_quelle": eaktz,
            "dgf_untervers": 1,
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte] != "")]
            else:
                if spalte == "dgf_untervers":
                    gefiltert[spalte] = pd.to_numeric(gefiltert[spalte], errors='coerce')
                gefiltert = gefiltert[gefiltert[spalte] == wert]
        wert_dgf = len(gefiltert)

        basis_wert = wert_hgf + wert_dgf
        return basis_wert

    def adressen_hinzunahme(gemeinde, gemarkung, eaktz, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "foerderazbnd": eaktz,
            "status_wechsel_pt": "In Antrag hinzunehmen",
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_entfernen(gemeinde, gemarkung, eaktz, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "foerderazbnd": eaktz,
            "status_wechsel_pt": "Aus Antrag entfernen",
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_in_bearbeitung(gemeinde, gemarkung, eaktz, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "status_wechsel_pt": "in Bearbeitung",
            "foerderazbnd": eaktz,
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte].astype(str).str.strip() != "")]
            elif isinstance(wert, list):
                gefiltert = gefiltert[gefiltert[spalte].isin(wert)]  # ← mehrere Werte
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    def adressen_aktuell(gemeinde, gemarkung, eaktz, df):

        # Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
        fixe_kriterien = {
            "gemark_nam": gemarkung,
            "status_wechsel_pt": "LEER",
            "foerderazbnd": eaktz,
        }

        """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
        gefiltert = df[df[SUCH_SPALTE_QUELLE] == gemeinde]

        for spalte, wert in fixe_kriterien.items():
            if wert == "NICHT_LEER":
                gefiltert = gefiltert[gefiltert[spalte].notna() & (gefiltert[spalte].astype(str).str.strip() != "")]
            elif wert == "LEER":
                gefiltert = gefiltert[gefiltert[spalte].isna() | (gefiltert[spalte].astype(str).str.strip() == "")]
            else:
                gefiltert = gefiltert[gefiltert[spalte] == wert]

        return len(gefiltert)

    # Basis-Excel laden
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Statusdatei: " + STATUS_DATEI + " wird geladen")
    wb = load_workbook(STATUS_DATEI)
    ws = wb[STATUS_SHEET]
    if data['Prod'] == "j":
        address_master = [home + "/" + data["oew_ablage_adressen"] + data["ADK_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["LBC_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["FDS_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["LRT_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["SIG_ADRESS_MASTER"],
                          home + "/" + data["oew_ablage_adressen"] + data["ZAK_ADRESS_MASTER"]
                          ]
    else:
        address_master = [home + "/" + data["oew_ablage_adressen"] + data["ADK_ADRESS_MASTER"]]
        #address_master = [home + "/" + data["oew_ablage_adressen"] + data["LBC_ADRESS_MASTER"]]
        #address_master = [home + "/" + data["oew_ablage_adressen"] + data["FDS_ADRESS_MASTER"]]

    for quell_datei in address_master:
        datum = datetime.today().strftime("%Y%m%d %H%M%S")
        print(datum + ": Adressdatei: " + quell_datei + " gestartet")

        if quell_datei == home + "/" + data["oew_ablage_adressen"] + data["ADK_ADRESS_MASTER"]:
            lk = "ADK"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["LBC_ADRESS_MASTER"]:
            lk = "LBC"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["FDS_ADRESS_MASTER"]:
            lk = "FDS"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["LRT_ADRESS_MASTER"]:
            lk = "LRT"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["SIG_ADRESS_MASTER"]:
            lk = "SIG"
        elif quell_datei == home + "/" + data["oew_ablage_adressen"] + data["ZAK_ADRESS_MASTER"]:
            lk = "ZAK"

        df = pd.read_excel(quell_datei, sheet_name=QUELL_SHEET, header=14)

        for row_idx in range(START_ROW, ws.max_row + 1):
            gemeinde = ws.cell(row=row_idx, column=4).value
            gemarkung = ws.cell(row=row_idx, column=5).value
            eaktz = ws.cell(row=row_idx, column=2).value
            landkreis = ws.cell(row=row_idx, column=3).value

            if landkreis and landkreis.startswith(lk):
                if gemeinde is None or str(gemeinde).strip() == "":
                    continue

                datum = datetime.today().strftime("%Y%m%d %H%M%S")
                print(datum + ": Auswertung Gemeinde : " + str(gemeinde))

                anzahl_basis_value = adressen_basis_treffer(gemeinde, gemarkung, eaktz, df)
                anzahl_hinzunahme_value = adressen_hinzunahme(gemeinde, gemarkung, eaktz, df)
                anzahl_entfernen_value = adressen_entfernen(gemeinde, gemarkung, eaktz, df)
                anzahl_in_bearbeitung_value = adressen_in_bearbeitung(gemeinde, gemarkung, eaktz, df)
                anzahl_aktuell_value = adressen_aktuell(gemeinde, gemarkung, eaktz, df)

                if anzahl_basis_value != 0:
                    ws.cell(row=row_idx, column=ADRESSEN_BASIS_COL, value=anzahl_basis_value)
                if anzahl_hinzunahme_value != 0:
                    ws.cell(row=row_idx, column=ADRESSEN_HINZUNAHME_COL, value=anzahl_hinzunahme_value)
                if anzahl_entfernen_value != 0:
                    ws.cell(row=row_idx, column=ADRESSEN_ENTFERNEN_COL, value=anzahl_entfernen_value)
                if anzahl_in_bearbeitung_value != 0:
                    ws.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG, value=anzahl_in_bearbeitung_value)
                if anzahl_aktuell_value != 0:
                    ws.cell(row=row_idx, column=ADRESSEN_AKTUELL, value=anzahl_aktuell_value)

    for row_idx in range(START_ROW, ws.max_row + 1):
        if ws.cell(row=row_idx, column=ADRESSEN_BASIS_COL).value is None:
            ws.cell(row=row_idx, column=ADRESSEN_BASIS_COL, value=0)
        if ws.cell(row=row_idx, column=ADRESSEN_HINZUNAHME_COL).value is None:
            ws.cell(row=row_idx, column=ADRESSEN_HINZUNAHME_COL, value=0)
        if ws.cell(row=row_idx, column=ADRESSEN_ENTFERNEN_COL).value is None:
            ws.cell(row=row_idx, column=ADRESSEN_ENTFERNEN_COL, value=0)
        if ws.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG).value is None:
            ws.cell(row=row_idx, column=ADRESSEN_IN_BEARBEITUNG, value=0)
        if ws.cell(row=row_idx, column=ADRESSEN_AKTUELL).value is None:
            ws.cell(row=row_idx, column=ADRESSEN_AKTUELL, value=0)

    wb.save(STATUS_DATEI)
    datum = datetime.today().strftime("%Y%m%d %H%M%S")
    print(datum + ": Adressmaster-Daten verarbeitet")


if __name__ == '__main__':
    main()
