import os
import json
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

home = os.path.expanduser('~')

my_dir = Path(__file__).parent
f = open(my_dir / 'config.json')
data = json.load(f)
input_dir = home + "/" + data['input_dir']
output_dir = home + "/" + data['output_dir']
f.close()

BASIS_DATEI  = input_dir + "AP24_Master_Status.xlsx"
QUELL_DATEI  = input_dir + "Masterliste.xlsx"
BASIS_SHEET  = "Master"
QUELL_SHEET  = "Adressliste_zur Bearbeitung_int"
START_ROW    = 5
ERGEBNIS_COL = 14  # Spalte N – hier wird die Anzahl reingeschrieben

# Quelldatei laden
df = pd.read_excel(QUELL_DATEI, sheet_name=QUELL_SHEET, header=12)

print(df.head())

# Fixe Kriterien definieren – Spaltenname (aus Header) und gewünschter Wert
fixe_kriterien = {
    "Förderprogramm":    "Hellgrau (HGF)",       # Spaltenname in Quelldatei : gesuchter Wert
    # weitere Kriterien einfach ergänzen...
}

# Dynamischer Suchwert kommt aus Spalte J der Basis → sucht in dieser Quellspalte
SUCH_SPALTE_QUELLE = "bauabschnitt_code"  # Spaltenname in der Quelldatei (entspricht Spalte B o.ä.)

def zaehle_treffer(suchwert):
    """Filtert den DataFrame nach allen Kriterien und gibt die Anzahl zurück."""
    gefiltert = df[df[SUCH_SPALTE_QUELLE] == suchwert]

    for spalte, wert in fixe_kriterien.items():
        gefiltert = gefiltert[gefiltert[spalte] == wert]

    return len(gefiltert)

# Basis-Excel laden
wb = load_workbook(BASIS_DATEI)
ws = wb[BASIS_SHEET]

for row_idx in range(START_ROW, ws.max_row + 1):
    suchwert = ws.cell(row=row_idx, column=10).value  # Spalte J = 10

    if suchwert is None:
        continue

    anzahl = zaehle_treffer(suchwert)
    ws.cell(row=row_idx, column=ERGEBNIS_COL, value=anzahl)

wb.save(BASIS_DATEI)
print("Fertig – Anzahlen wurden geschrieben.")